import os
import shutil
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta

def arquivos_zip(path, pathdestino):
    pasta_download = os.path.expanduser("~/Downloads")
    pasta_destino_zip = r"digite sua pasta"

    arquivos_zip_encontrados = [z for z in os.listdir(pasta_download) if z.endswith(".zip") and "ccr-tickets" in z.lower()]

    if not arquivos_zip_encontrados:
        print("Nenhum arquivo .zip com 'ccr-tickets' encontrado na pasta de downloads.")
    else:
        if os.path.exists(pasta_destino_zip):
            arquivos_destino_zip = [z for z in os.listdir(pasta_destino_zip) if "ccr-tickets" in z.lower()]
            for arquivo in arquivos_destino_zip:
                remover_arquivo = os.path.join(pasta_destino_zip, arquivo)
                os.remove(remover_arquivo)
                print(f"[ZIP] Arquivo removido do destino: {remover_arquivo}")

            for zip_nome in arquivos_zip_encontrados:
                caminho_zip = os.path.join(pasta_download, zip_nome)
                with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
                    for nome_arquivo in zip_ref.namelist():
                        if "ccr-tickets" in nome_arquivo.lower():
                            zip_ref.extract(nome_arquivo, pasta_destino_zip)
                            print(f"[ZIP] Arquivo extraído: {nome_arquivo}")
                print(f"[ZIP] Arquivo ZIP processado: {zip_nome}")
        else:
            print(f"A pasta de destino para ZIPs não existe: {pasta_destino_zip}")

def arquivo_mkt(path, pathdestino):
    pasta_download = os.path.expanduser("~/Downloads")  
    pasta_destino_motivos = pathdestino 

    arquivos_motivos = [f for f in os.listdir(pasta_download) if "Motivos de contato" in f]

    if not arquivos_motivos:
        print("Nenhum arquivo com 'Motivos de contato' encontrado na pasta de downloads.")
    else:
        if os.path.exists(pasta_destino_motivos):
            arquivos_destino_motivos = [f for f in os.listdir(pasta_destino_motivos) if "motivos" in f.lower()]
            for arquivo in arquivos_destino_motivos:
                caminho_remover = os.path.join(pasta_destino_motivos, arquivo)
                os.remove(caminho_remover)
                print(f"[MOTIVOS] Arquivo removido do destino: {caminho_remover}")
        else:
            print(f"A pasta de destino para 'Motivos de contato' não existe: {pasta_destino_motivos}")
            return 

        for arquivo in arquivos_motivos:
            caminho_arquivo = os.path.join(pasta_download, arquivo)
            caminho_final = os.path.join(pasta_destino_motivos, arquivo)

            if os.path.exists(caminho_arquivo):
                shutil.move(caminho_arquivo, caminho_final)
                print(f"[MOTIVOS] Arquivo movido: {caminho_final}")
            else:
                print(f"[MOTIVOS] Arquivo não encontrado no momento da cópia: {caminho_arquivo}")

def month_offset(offset_meses):
    data_ajustada = datetime.now() + relativedelta(months=offset_meses)
    return data_ajustada.strftime('%m')

# def atualizar_power_query(wb, nome_aba):
#     ws = wb[nome_aba]
#     ultima_linha = ws.max_row

#     for col in range(18, 23):  # Colunas R (18) até V (22)
#         letra_coluna = get_column_letter(col)
#         celula_formula = f"{letra_coluna}2"
#         formula = ws[celula_formula].value

#         if formula and str(formula).startswith("="):
#             # Replicar fórmula da linha 2 para baixo
#             for linha in range(3, ultima_linha + 1):
#                 ws[f"{letra_coluna}{linha}"].value = formula

#             # Converter da linha 3 em diante para valores
#             for linha in range(3, ultima_linha + 1):
#                 valor = ws[f"{letra_coluna}{linha}"].value
#                 ws[f"{letra_coluna}{linha}"].value = valor
#                 ws[f"{letra_coluna}{linha}"].data_type = 'n'
#         else:
#             print(f"Nenhuma fórmula encontrada em {celula_formula}")

# Caminho do arquivo com offset de mês
# offset = month_offset(0)
# archive_base_pesquisa = fr"\\riachu-ccfs\callnt-files\CCR - Esquadrões\5. Relatórios\MARKETPLACE\MIS\Base Pesquisas\2025{offset} - PESQUISA MARKETPLACE.xlsx"
# # Carregar e atualizar planilha
# wb = load_workbook(archive_base_pesquisa)
# atualizar_power_query(wb, nome_aba="ConsultaPesquisa")  # ajuste o nome da aba conforme necessário
# wb.save(archive_base_pesquisa)

# Executar etapas de arquivos
arquivos_zip(
    path="C:\\Users\\4061014\\Downloads",
    pathdestino="\\\\riachu-ccfs\\callnt-files\\CCR - Esquadrões\\5. Relatórios\\MARKETPLACE\\MIS\\Base Cockpit"
)        

arquivo_mkt(
    path="C:\\Users\\4061014\\Downloads", 
    pathdestino="\\\\riachu-ccfs\\callnt-files\\Planejamento\\02_MIS\\01_Relatórios\\E-Commerce\\Bases\\Bases Contact Rate\\Base Dynamics"
)

print("Processamento geral concluído com sucesso")